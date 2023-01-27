import pyautogui
import pyperclip
import time
import csv
from datetime import datetime
from openpyxl import load_workbook

load_wb = load_workbook("C:\\시스템관리팀_전표_자동화.xlsx", data_only=True)
sheet_name = str(datetime.now().year)[2:] + "년 " + str(datetime.now().strftime('%m')) + "월"  # 예) 23년 01월
load_ws = load_wb[sheet_name]  # 시트 이름으로 불러오기

company, amount, service, gl_date, cut_off, dr_account, tax_code, description, vendor = [], [], [], [], [], [], [], [], []


class InvoiceEntry:
    option = ""

    def __init__(self, option):
        InvoiceEntry.option = option

    @staticmethod
    def activate_window():
        win = pyautogui.getWindowsWithTitle('심텍 가동계')[0]
        if win.isMinimized:
            win.restore()
        win.activate()  # 윈도우(창) 활성화


class ManipulateKeyboardMouse:
    @staticmethod
    def move_and_click(x, y):
        pyautogui.moveTo(x, y)
        pyautogui.click()

    @staticmethod
    def press_key(key, n=1):
        for _ in range(n):
            pyautogui.keyDown(key)
            pyautogui.keyUp(key)

    @staticmethod
    def press_keys(*keys):  # 기본값 1로 설정
        for key in list(keys):
            pyautogui.keyDown(key)
            pyautogui.keyUp(key)

    @staticmethod
    def copy_and_paste_to_oracle(key):
        pyperclip.copy(key)
        pyautogui.hotkey("ctrl", "v")


class ManipulateData:
    @staticmethod
    def conv_gl_date(date):  # 엑셀 표기 형식의 날짜를 YYYY-MM-DD 형태로 변환한다.
        if date is None:
            return datetime.today().strftime('%Y-%m-%d')
        else:
            return str(date.year) + "-" + str(date.month) + "-" + str(date.day)


class SetTaxAccountRule:
    @staticmethod
    def tax_code_rule(service_name):
        # M365 전표 세금 코드
        if InvoiceEntry.option == "m365":
            if service_name in ["M365 STG", "M365 STP", "M365 STX", "M365 T.E TECH"]:
                return "C_IN_NONREC_10%"
            else:
                return "C_IN_TAX_10%"
        # 통신 전표 세금 코드
        elif InvoiceEntry.option == "communication":
            if "연수원" in service_name:
                return "C2_IN_TAX_10%"
            elif "SKT" in service_name:
                return "C_IN_NONREC_10%"
            else:
                return "O2_IN_TAX_10%" if "오창" in service_name else "C_IN_TAX_10%"
        # 기타 전표 세금 코드
        else:
            return "C_IN_TAX_10%"


class ImportExportData:
    @staticmethod
    def import_data():
        for datas_idx, row in enumerate(load_ws['J4': "J" + str(load_wb.active.max_row)], start=4):
            for _ in row:
                # 청구 금액이 없는 경우 또는 내용이 없는 경우 무시
                if load_ws['D' + str(datas_idx)].value is None or load_ws['B' + str(datas_idx)].value is None:
                    continue

                company.append(load_ws['J' + str(datas_idx)].value)  # Site Name 데이터 입력
                amount.append(str(load_ws['D' + str(datas_idx)].value))  # 청구 금액 데이터 입력
                service.append(load_ws['B' + str(datas_idx)].value)  # 내용 데이터 입력
                gl_date.append(ManipulateData.conv_gl_date(load_ws['I' + str(datas_idx)].value))  # 품의 일자 데이터 입력
                cut_off.append(str(load_ws['G' + str(datas_idx)].value))  # 잡이익 데이터 입력
                dr_account.append("1.000.132000." + str(load_ws['K' + str(datas_idx)].value) + ".0000.000000.0.000")  # Dr.Account 데이터 입력
                tax_code.append(SetTaxAccountRule.tax_code_rule(load_ws['B' + str(datas_idx)].value))  # Tax Code 데이터 입력
                vendor.append(load_ws['A' + str(datas_idx)].value if load_ws['A' + str(datas_idx)].value is not None else vendor[-1])  # 거래처 데이터 입력
                description.append(str(datetime.now().month) + "월_" + vendor[-1] + "(" + service[-1] + ")")  # Description 데이터 입력

    @staticmethod
    def import_header():
        ManipulateKeyboardMouse.move_and_click(153, 253)  # 첫번째 행 위치 클릭
        for i in range(len(company)):
            if i != 0:
                pyautogui.keyDown('ctrl')
                for j in range(i):
                    pyautogui.press('down')
                ManipulateKeyboardMouse.press_keys('ctrl', 'tab', 'tab')
            ManipulateKeyboardMouse.press_keys('s', 'tab', 'tab')
            ManipulateKeyboardMouse.copy_and_paste_to_oracle(company[i].replace("㈜", "(주)"))  # Site Name 입력 (특수 문자 ㈜ 처리)
            ManipulateKeyboardMouse.press_key('\t', 2)
            ManipulateKeyboardMouse.copy_and_paste_to_oracle(gl_date[i])  # GL Date 품의 일자 입력
            if InvoiceEntry.option == "etc":  # 기타 전표는 Cr.Account : 1.000.000000.210300100.0000.000000.0.000 로 고정
                ManipulateKeyboardMouse.press_key('\t', 6)
                ManipulateKeyboardMouse.copy_and_paste_to_oracle("1.000.000000.210300100.0000.000000.0.000")
                ManipulateKeyboardMouse.press_key('\t', 2)
            else:
                ManipulateKeyboardMouse.press_key('\t', 8)
            ManipulateKeyboardMouse.copy_and_paste_to_oracle(description[i])
            ManipulateKeyboardMouse.press_key('\t', 3)
            ManipulateKeyboardMouse.copy_and_paste_to_oracle('cash')
            ManipulateKeyboardMouse.press_key('\t', 4)
            ManipulateKeyboardMouse.press_keys('v', 'tab', 'e', 'enter')
            if i == len(company) - 1:  # record 추가 시
                pyautogui.hotkey("ctrl", "down")
                ManipulateKeyboardMouse.press_keys('tab', 'tab')
        pyautogui.hotkey("ctrl", "s")

    @staticmethod
    def import_vat():
        items_in_one_scroll = 8
        ManipulateKeyboardMouse.press_key("pgup", len(company) // 7 + 2)  # 최상단으로 Scroll Up

        for i in range(len(company)):
            if i < items_in_one_scroll:
                ManipulateKeyboardMouse.move_and_click(354, 252 + (i * 24))
            else:
                ManipulateKeyboardMouse.move_and_click(354, 252 + (7 * 24))
                pyautogui.press("down")

            ManipulateKeyboardMouse.move_and_click(233, 534)
            ManipulateKeyboardMouse.copy_and_paste_to_oracle(dr_account[i])  # Dr.Account 입력
            ManipulateKeyboardMouse.press_key('\t', 2)
            ManipulateKeyboardMouse.copy_and_paste_to_oracle(amount[i])  # Amount 입력
            ManipulateKeyboardMouse.press_key('\t', 2)
            ManipulateKeyboardMouse.copy_and_paste_to_oracle(tax_code[i])  # Tax Code 입력
            ManipulateKeyboardMouse.press_key('\t')
            pyautogui.press("esc")
            pyautogui.hotkey("ctrl", "s")  # save 버튼
            ManipulateKeyboardMouse.move_and_click(124, 695)  # Calculate Tax 버튼
            pyautogui.hotkey("ctrl", "s")  # save 버튼
            # 절사액 입력 부분
            if cut_off[i] != 0 and InvoiceEntry.option == "communication" and company[i] != "INPOBANGK":
                ManipulateKeyboardMouse.move_and_click(233, 580)  # 하단 Dr.Account 위치 이동
                time.sleep(1)
                pyautogui.click()
                ManipulateKeyboardMouse.copy_and_paste_to_oracle("1.000.132000.420001700.0000.000000.0.000")
                ManipulateKeyboardMouse.press_key('\t', 2)
                ManipulateKeyboardMouse.copy_and_paste_to_oracle(str(-int(cut_off[i])))
                pyautogui.hotkey("ctrl", "s")  # save 버튼

    @staticmethod
    def export_to_csv():
        export_items = []
        for pairs in zip(company, amount, service, gl_date, cut_off, dr_account, tax_code, description):
            export_items.append(list(pairs))
            print(pairs)

        with open("C:\\" + str(datetime.now().month) + "월 시스템관리팀 " + InvoiceEntry.option + " 전표 입력 항목.csv", 'w', newline='\n') as f:
            writer = csv.DictWriter(f, fieldnames=["거래처", "청구 금액", "내용", "품의 일자", "잡이익", "Dr.Account", "Tax Code", "Description"])
            writer.writeheader()
            write = csv.writer(f)
            write.writerows(export_items)
