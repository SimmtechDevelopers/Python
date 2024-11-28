from datetime import datetime
from openpyxl import load_workbook
from decimal  import Decimal
import sys
import os

def STG_currency():
# 엑셀 파일 경로


    exe_path = sys.argv[0]
    #현재 스크립트가 위치한 디렉토리 경로를 얻습니다.
    # script_dir = os.path.dirname(exe_path)

    file_name = 'レート集計表.xlsx'
    # excel_file_path = os.path.join(script_dir, file_name)

    ## file_watcher 용 경로 지정
    excel_file_path = os.path.join("\\\\172.16.1.12\\PrevailingRate", file_name)

    print(excel_file_path)

    # 현재 날짜 정보 가져오기 (년, 월, 일)
    today_str = str(datetime.now().strftime("%Y-%m-%d"))
    today =datetime.now()
    year = today.year
    month = today.month
    day = today.day

    # 오늘 날짜에 해당하는 시트 이름 생성
    sheet_name = f"{year}.{month}"+"月"  # ex) 2024.03

    # 엑셀 파일 로드
    workbook = load_workbook(filename=excel_file_path)
    currency_rate = []

    result ={}

    # 시트가 존재하는지 확인하고 데이터 읽기
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        found_row = None
        for row in sheet.iter_rows(min_row=3,values_only=True):
            date_obj = str(row[1])
            formatted_date = date_obj[:10]
            if formatted_date == today_str:
                result["USD"] = row[2]
                result["MYR"] = row[3]
                result["KRW"] = row[4]
                break

    return result

# print(test)
